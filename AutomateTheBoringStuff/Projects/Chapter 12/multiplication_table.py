#! python3
# multiplication_table.py - Creates a multiplication table in excel with N rows

"""
Create a program multiplicationTable.py that takes a number N from the command line
and creates an N×N multiplication table in an Excel spreadsheet.

Pandas has been more popular with the work that I will be working on so I'm just going over the module
without a huge focus on it later on but good to know about the functionality
"""

import openpyxl
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

os.chdir(r'')

# sys.argv[1] will be the first argument entered
# sys.argv[0] is the first argument and it's a string containing the filename
n_table = 10  # int(sys.argv[1])

# Open a new workbook and use the active sheet
wb = openpyxl.Workbook()
sheet = wb.active

"""
Start from row 2 since A1 is empty. This will be the first loop and will write out the rows and columns
up to the number that was entered in the command line. When trying to add a value to a cell you need a letter
and number, but in a string format.

This creates the outer layer of the numbers
"""
for r in range(2, n_table+2):
    sheet['A' + str(r)] = r-1
    col_let = get_column_letter(r)
    sheet[col_let + str(1)] = r-1
    # Set to bold
    sheet['A' + str(r)].font = Font(bold=True)
    sheet[col_let + str(1)].font = Font(bold=True)

"""
Loop to create the values inside the box created above
"""
for col in range(2, n_table+2):
    for row in range(2, n_table+2):
        col_letters = get_column_letter(col)
        sheet[col_letters + str(row)] = (col-1) * (row-1)
        print(str(col) + '-' + str(row))

wb.save('multiplication_table.xlsx')