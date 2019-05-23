#! python3
# excel_to_csv.py - Converts all Excel files in a working directory into csv files

import openpyxl
import csv
import os
import re
import shutil

"""
This first block is for setting up. The idea is to create two new directories so you are working with clean
areas. However you can edit these settings to your personal preference
"""

# Set working directory
path = os.getcwd()
print(path)

"""
Create a new one with selective_copy.py and import a bunch of excel spreadsheets.
Create a new variable in this directory for use later with shutil
Create a new directory inside the created directory. Store new CSV files here
"""
os.makedirs(path + '\\' + 'csv_files', exist_ok=True)

# Create a new variable of this directory to use later with shutil (in conjunction with path)
new_path = path + '\\' + 'csv_files'

"""
Next block is reading the excel file, creating the new CSV files, then copying the excel files into the new csv files.
"""
for excel_file in os.listdir('.'):
    # Skip the non-xlsx files, load the workbook object
    if not excel_file.endswith('.xlsx'):
        continue
    workbook = openpyxl.load_workbook(excel_file)
    # Loop through every sheet in the workbook
    for sheets in workbook.sheetnames:
        wb_name = re.sub('.xlsx', '', excel_file)  # remove the extension to clean it up
        csv_name = wb_name + '_' + sheets + '.csv'  # newly created file

        # Open CSV file and create your writer object
        csv_file = open(csv_name, 'w', newline='')
        csv_writer = csv.writer(csv_file)
        sheet = workbook.active

        # The first part of the loop to create a new list for each row in each sheet
        for row_num in range(1, sheet.max_row + 1):
            row_data = []
            # Use the list created above to append each cell value through each column
            for col_num in range(1, sheet.max_column + 1):
                cell_data = sheet.cell(row=row_num, column=col_num).value
                row_data.append(cell_data)

            # This is the writing step, only write once per row
            csv_writer.writerow(row_data)

            # Close and move the file
        csv_file.close()
        shutil.move(os.path.join(path, csv_name), os.path.join(new_path, csv_name))

print('Finished')